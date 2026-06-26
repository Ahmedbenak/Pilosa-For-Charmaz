#!/usr/bin/env python3
"""
gerer_tracabilite.py — Gestionnaire du classeur de traçabilité Charmaz
Chemin stable : methodologie/tracabilite.xlsx

Dépendances : openpyxl
    pip install openpyxl

Commandes :
    python methodologie/gerer_tracabilite.py --init     # Créer le classeur vide
    python methodologie/gerer_tracabilite.py --check    # Reconstruire checkpoint depuis le classeur
    python methodologie/gerer_tracabilite.py --status   # Afficher l'état courant
    python methodologie/gerer_tracabilite.py --help     # Aide
"""

import argparse
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERREUR : openpyxl non installé. Exécuter : pip install openpyxl")
    sys.exit(1)

# ── Chemins stables ────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
XLSX_PATH = SCRIPT_DIR / "tracabilite.xlsx"
XLSX_TMP = SCRIPT_DIR / "tracabilite.xlsx.tmp"
CHECKPOINT_PATH = SCRIPT_DIR / "checkpoint.json"
ETAT_PATH = SCRIPT_DIR / "ETAT_AVANCEMENT.md"

# ── Définition des 11 onglets avec leurs colonnes ─────────────────────────────
ONGLETS = {
    "00_Corpus_Source": [
        "id_entretien", "fichier_source", "format", "date_entretien",
        "duree_approx", "nb_segments", "notes_non_identifiantes"
    ],
    "01_Segments": [
        "id_segment", "id_entretien", "texte_verbatim", "contexte",
        "locuteur", "position_ligne", "position_paragraphe"
    ],
    "02_Codage_Initial": [
        "id_code", "id_segment", "texte_code", "in_vivo",
        "question_p38_mobilisee", "question_p39_mobilisee",
        "processus_amorce", "note_analytique",
        "regle_regroupement"
    ],
    "03_Codage_Focalise": [
        "id_code_focalise", "intitule_focalise", "liste_id_code",
        "frequence", "portee_corpus", "processus_amorce_synthetise",
        "regle_regroupement"
    ],
    "04_Categories": [
        "id_categorie", "nom_categorie", "type", "liste_id_code_focalise",
        "proprietes", "conditions_apparition", "conditions_maintien",
        "conditions_changement", "consequences", "relations_autres_categories",
        "processus_generique", "ancrage_empirique",
        "test_sensitizing_concept", "verbatim_illustratif",
        "regle_regroupement"
    ],
    "05_Comparaisons": [
        "id_comparaison", "type_comparaison", "elements_compares",
        "question_posee", "resultat", "impact_sur_categorie",
        "refs_id_categorie_modifiees"
    ],
    "06_Memos": [
        "id_memo", "titre", "id_categorie_principale",
        "id_categories_secondaires", "id_segments_cites",
        "fichier_memo", "questions_ouvertes",
        "regle_regroupement"
    ],
    "07_Echantillonnage_Theorique": [
        "id_echantillon", "id_categorie", "propriete_developpee",
        "id_segment_relus", "resultat_saturation", "limite_signalee"
    ],
    "08_Construction_Themes": [
        "id_theme", "id_sous_theme", "intitule_theme", "regle_regroupement",
        "id_categorie_sources", "justification_regroupement",
        "processus_generique"
    ],
    "09_Theorie_Finale": [
        "id_enonce", "texte_enonce", "id_theme_sources",
        "processus_generique_exprime", "ancrage_ultimate",
        "regle_regroupement"
    ],
    "10_Journal_Tracabilite": [
        "id_decision", "agent", "horodatage", "type_operation",
        "id_element_verifie", "entrees", "sorties",
        "justification", "violations", "resultat", "action_requise"
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# ── Écriture atomique ──────────────────────────────────────────────────────────

def ecrire_atomique(wb: "openpyxl.Workbook"):
    """Sauvegarde le classeur de façon atomique : .tmp → replace."""
    wb.save(XLSX_TMP)
    shutil.move(str(XLSX_TMP), str(XLSX_PATH))


# ── Création du classeur vide ──────────────────────────────────────────────────

def creer_classeur():
    """Crée tracabilite.xlsx avec les 11 onglets vides et leurs en-têtes."""
    if XLSX_PATH.exists():
        print(f"INFO : {XLSX_PATH} existe déjà — non écrasé.")
        print("       Utilisez --check pour reconstruire le checkpoint depuis le classeur existant.")
        return

    wb = openpyxl.Workbook()
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, columns in ONGLETS.items():
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    ecrire_atomique(wb)
    print(f"OK : {XLSX_PATH} créé avec {len(ONGLETS)} onglets.")

    # Initialiser le checkpoint si absent
    if not CHECKPOINT_PATH.exists():
        initialiser_checkpoint()


# ── Lecture des ids existants ──────────────────────────────────────────────────

def lire_derniers_ids(onglet: str) -> set:
    """Retourne l'ensemble des valeurs de la première colonne (ids) d'un onglet."""
    if not XLSX_PATH.exists():
        return set()
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if onglet not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[onglet]
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ids.add(str(row[0]))
    wb.close()
    return ids


def lire_dernier_id_sequence(onglet: str) -> str | None:
    """Retourne le dernier id présent dans un onglet (dernière ligne non vide, col 0)."""
    if not XLSX_PATH.exists():
        return None
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if onglet not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[onglet]
    last_id = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            last_id = str(row[0])
    wb.close()
    return last_id


# ── Écriture de lignes avec idempotence ───────────────────────────────────────

def ecrire_lignes_atomique(onglet: str, lignes: list[list]) -> int:
    """
    Écrit les lignes dans l'onglet donné, en sautant celles dont l'id existe déjà.
    Écriture atomique : .tmp → replace.
    Retourne le nombre de lignes effectivement écrites.
    """
    if not XLSX_PATH.exists():
        print(f"ERREUR : {XLSX_PATH} n'existe pas. Lancer --init d'abord.")
        return 0

    ids_existants = lire_derniers_ids(onglet)
    lignes_nouvelles = [l for l in lignes if str(l[0]) not in ids_existants]

    if not lignes_nouvelles:
        return 0

    wb = openpyxl.load_workbook(XLSX_PATH)
    if onglet not in wb.sheetnames:
        print(f"ERREUR : onglet '{onglet}' introuvable dans le classeur.")
        return 0

    ws = wb[onglet]
    for ligne in lignes_nouvelles:
        ws.append(ligne)

    ecrire_atomique(wb)
    return len(lignes_nouvelles)


# ── Checkpoint ────────────────────────────────────────────────────────────────

def initialiser_checkpoint():
    """Écrit checkpoint.json avec statut a_faire si absent."""
    if CHECKPOINT_PATH.exists():
        return
    checkpoint = _checkpoint_vide()
    _ecrire_checkpoint(checkpoint)
    print(f"OK : {CHECKPOINT_PATH} initialisé.")


def lire_checkpoint() -> dict:
    if not CHECKPOINT_PATH.exists():
        return _checkpoint_vide()
    with open(CHECKPOINT_PATH, encoding="utf-8") as f:
        return json.load(f)


def _checkpoint_vide() -> dict:
    now = _now()
    return {
        "_schema": "charmaz-checkpoint-v1",
        "statut": "a_faire",
        "premier_lancement": None,
        "dernier_lancement": None,
        "dernier_heartbeat": None,
        "dernier_run_id": None,
        "corpus": {"total_entretiens": 0, "entretiens": []},
        "etapes": {
            e: {"statut": "a_faire", "dernier_id_valide": None, "nb_valides": 0}
            for e in [
                "segmentation", "codage_initial", "codage_focalise",
                "categorisation", "comparaison_constante", "memos",
                "echantillonnage_theorique", "construction_themes",
                "theorie_finale", "verification"
            ]
        },
        "runs": [],
    }


def _ecrire_checkpoint(checkpoint: dict):
    tmp = CHECKPOINT_PATH.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, ensure_ascii=False, indent=2)
    shutil.move(str(tmp), str(CHECKPOINT_PATH))


def mettre_a_jour_checkpoint(etape: str, dernier_id: str, nb_valides_delta: int = 1):
    """Met à jour le checkpoint après écriture validée d'un identifiant."""
    checkpoint = lire_checkpoint()
    checkpoint["dernier_heartbeat"] = _now()
    if etape in checkpoint["etapes"]:
        checkpoint["etapes"][etape]["dernier_id_valide"] = dernier_id
        checkpoint["etapes"][etape]["nb_valides"] = (
            checkpoint["etapes"][etape].get("nb_valides", 0) + nb_valides_delta
        )
        checkpoint["etapes"][etape]["statut"] = "en_cours"
    if checkpoint["statut"] == "a_faire":
        checkpoint["statut"] = "en_cours"
    _ecrire_checkpoint(checkpoint)


def marquer_termine():
    """Seul l'Agent Vérification appelle cette fonction après certification complète."""
    checkpoint = lire_checkpoint()
    checkpoint["statut"] = "termine"
    checkpoint["dernier_heartbeat"] = _now()
    checkpoint["etapes"]["verification"]["statut"] = "termine"
    checkpoint["etapes"]["verification"]["certifie_complet"] = True
    _ecrire_checkpoint(checkpoint)
    regenerer_etat_avancement()
    print("OK : analyse marquée TERMINÉE dans checkpoint.json.")


# ── Reconstruction du checkpoint depuis le classeur ───────────────────────────

ONGLET_ETAPE_MAP = {
    "01_Segments": "segmentation",
    "02_Codage_Initial": "codage_initial",
    "03_Codage_Focalise": "codage_focalise",
    "04_Categories": "categorisation",
    "05_Comparaisons": "comparaison_constante",
    "06_Memos": "memos",
    "07_Echantillonnage_Theorique": "echantillonnage_theorique",
    "08_Construction_Themes": "construction_themes",
    "09_Theorie_Finale": "theorie_finale",
    "10_Journal_Tracabilite": "verification",
}


def reconstruire_checkpoint():
    """
    Relit le classeur pour reconstruire l'état réel de chaque étape.
    Utile après une interruption brutale où le checkpoint est désynchronisé.
    """
    if not XLSX_PATH.exists():
        print(f"ERREUR : {XLSX_PATH} introuvable — lancer --init d'abord.")
        return

    checkpoint = lire_checkpoint()
    modifie = False

    for onglet, etape in ONGLET_ETAPE_MAP.items():
        dernier_id = lire_dernier_id_sequence(onglet)
        ids_set = lire_derniers_ids(onglet)
        nb = len(ids_set)
        if dernier_id != checkpoint["etapes"][etape].get("dernier_id_valide") or \
           nb != checkpoint["etapes"][etape].get("nb_valides", 0):
            checkpoint["etapes"][etape]["dernier_id_valide"] = dernier_id
            checkpoint["etapes"][etape]["nb_valides"] = nb
            if dernier_id and checkpoint["etapes"][etape]["statut"] == "a_faire":
                checkpoint["etapes"][etape]["statut"] = "interrompu"
            modifie = True
            print(f"  {etape}: dernier_id={dernier_id}, nb={nb}")

    if modifie:
        if checkpoint["statut"] == "a_faire":
            checkpoint["statut"] = "interrompu"
        checkpoint["dernier_heartbeat"] = _now()
        _ecrire_checkpoint(checkpoint)
        print(f"OK : checkpoint.json reconstruit depuis {XLSX_PATH}.")
    else:
        print("INFO : checkpoint.json déjà synchronisé avec le classeur.")

    regenerer_etat_avancement()


# ── Régénération de ETAT_AVANCEMENT.md ───────────────────────────────────────

def regenerer_etat_avancement():
    checkpoint = lire_checkpoint()
    now = _now()

    etapes_labels = {
        "segmentation": "1. Segmentation",
        "codage_initial": "2. Codage initial",
        "codage_focalise": "3. Codage focalisé",
        "categorisation": "4. Catégorisation",
        "comparaison_constante": "5. Comparaison constante",
        "memos": "6. Mémos",
        "echantillonnage_theorique": "7. Échantillonnage théorique",
        "construction_themes": "8. Construction thèmes",
        "theorie_finale": "9. Théorie finale",
        "verification": "10. Vérification & Traçabilité",
    }

    # Trouver étape en cours
    etape_courante = "—"
    for etape_key, val in checkpoint["etapes"].items():
        if val["statut"] in ("en_cours", "interrompu"):
            etape_courante = etapes_labels.get(etape_key, etape_key)
            break

    # Point de reprise
    point_reprise = "—"
    for etape_key, val in checkpoint["etapes"].items():
        if val["statut"] in ("en_cours", "interrompu") and val.get("dernier_id_valide"):
            point_reprise = f"{etapes_labels[etape_key]} — après {val['dernier_id_valide']}"
            break

    # Progression entretiens
    entretiens = checkpoint["corpus"].get("entretiens", [])
    total = checkpoint["corpus"].get("total_entretiens", 0)
    termines = sum(1 for e in entretiens if e.get("statut") == "termine")
    progression = f"{termines} / {total if total else '?'} entretiens"

    # Tableau étapes
    lignes_etapes = []
    for etape_key, label in etapes_labels.items():
        val = checkpoint["etapes"][etape_key]
        statut = f"`{val['statut']}`"
        dernier = val.get("dernier_id_valide") or "—"
        nb = val.get("nb_valides", 0)
        lignes_etapes.append(f"| {label} | {statut} | {dernier} | {nb} |")

    # Tableau entretiens
    if entretiens:
        lignes_ent = ["| id_entretien | fichier | statut | étape en cours | dernier_segment_valide |",
                      "|---|---|---|---|---|"]
        for e in entretiens:
            lignes_ent.append(
                f"| {e.get('id_entretien','—')} | {e.get('fichier','—')} | "
                f"`{e.get('statut','—')}` | {e.get('etape_en_cours','—')} | "
                f"{e.get('dernier_segment_valide','—')} |"
            )
        section_entretiens = "\n".join(lignes_ent)
    else:
        section_entretiens = "_Aucun entretien traité pour l'instant._"

    # Historique des sessions (runs)
    runs = checkpoint.get("runs", [])
    if runs:
        lignes_runs = ["| run_id | démarrage | arrêt | cause_arret | nb_ids_écrits |",
                       "|---|---|---|---|---|"]
        for r in runs[-10:]:  # 10 derniers runs
            lignes_runs.append(
                f"| {r.get('run_id','—')} | {r.get('debut','—')} | {r.get('fin','—')} | "
                f"{r.get('cause_arret','—')} | {r.get('nb_ids_ecrits','—')} |"
            )
        section_runs = "\n".join(lignes_runs)
    else:
        section_runs = "_Aucune session enregistrée._"

    contenu = f"""---
type: etat_avancement
generated_by: charmaz-checkpoint
statut: {checkpoint['statut']}
mise_a_jour: {now}
---

# État d'avancement — Analyse Charmaz

> Ce fichier est régénéré automatiquement par l'orchestrateur Charmaz à chaque lancement et après chaque mise à jour du checkpoint. Ne pas modifier manuellement.
> Source de vérité machine : `methodologie/checkpoint.json`

---

## Résumé

| Champ | Valeur |
|---|---|
| **Statut global** | `{checkpoint['statut']}` |
| Premier lancement | {checkpoint.get('premier_lancement') or '—'} |
| Dernier lancement | {checkpoint.get('dernier_lancement') or '—'} |
| Dernier heartbeat | {checkpoint.get('dernier_heartbeat') or '—'} |
| **Progression entretiens** | {progression} |
| **Étape Charmaz en cours** | {etape_courante} |
| **Point de reprise exact** | {point_reprise} |

## Commande de reprise

```
# L'orchestrateur Charmaz lit automatiquement checkpoint.json au démarrage.
# Aucune commande spéciale — relancer simplement la pipeline Charmaz.
# Elle détectera le statut "{checkpoint['statut']}" et reprendra au dernier id validé.
```

---

## Détail par entretien

{section_entretiens}

---

## Détail par étape Charmaz

| Étape | Statut | Dernier id validé | Nb validés |
|---|---|---|---|
{chr(10).join(lignes_etapes)}

---

## Historique des sessions

{section_runs}

---

## Notes sur les interruptions

- Tout arrêt sans marqueur explicite de fin (`termine`) laisse le statut à `interrompu`.
- La dernière unité traitée avant l'interruption peut être incomplète — elle est retraitée au prochain lancement, jamais sautée.
- Le statut `termine` n'est posé que par l'Agent Vérification & Traçabilité après contrôle complet du chaînage.
"""

    with open(ETAT_PATH, "w", encoding="utf-8") as f:
        f.write(contenu)


# ── Affichage du statut ───────────────────────────────────────────────────────

def afficher_statut():
    checkpoint = lire_checkpoint()
    print(f"\n=== Statut Charmaz ===")
    print(f"  Statut global   : {checkpoint['statut']}")
    print(f"  Dernier heartbeat : {checkpoint.get('dernier_heartbeat') or '—'}")
    print(f"  Corpus : {checkpoint['corpus'].get('total_entretiens', 0)} entretiens")
    print()
    for etape, val in checkpoint["etapes"].items():
        status = val["statut"]
        last_id = val.get("dernier_id_valide") or "—"
        nb = val.get("nb_valides", 0)
        print(f"  {etape:35s} {status:12s} dernier={last_id} nb={nb}")
    print()


# ── Lecture d'un onglet vers JSON (pour agents) ───────────────────────────────

def lire_onglet_json(onglet: str, limit: int = 0, offset: int = 0) -> list[dict]:
    """Lit un onglet et retourne une liste de dicts (clé=valeur)."""
    if not XLSX_PATH.exists():
        return []
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if onglet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[onglet]
    headers = None
    rows = []
    count = 0
    skipped = 0
    for row in ws.iter_rows(values_only=True):
        vals = [str(c) if c is not None else "" for c in row]
        if headers is None:
            headers = vals
            continue
        if not vals[0]:
            continue
        skipped += 1
        if skipped <= offset:
            continue
        rows.append(dict(zip(headers, vals)))
        count += 1
        if limit > 0 and count >= limit:
            break
    wb.close()
    return rows


def ecrire_ligne_json(onglet: str, data: dict) -> int:
    """Écrit une ligne dans un onglet depuis un dict {colonne: valeur}.
    Retourne 1 si écrit, 0 si ignoré (id déjà présent)."""
    cols = ONGLETS.get(onglet, [])
    if not cols:
        print(f"ERREUR: onglet '{onglet}' inconnu")
        return 0
    ligne = [data.get(c, "") for c in cols]
    return ecrire_lignes_atomique(onglet, [ligne])


def ecrire_lignes_depuis_json(onglet: str, lignes: list[dict]) -> int:
    """Écrit plusieurs lignes depuis une liste de dicts {colonne: valeur}."""
    cols = ONGLETS.get(onglet, [])
    if not cols:
        return 0
    data = [[l.get(c, "") for c in cols] for l in lignes]
    return ecrire_lignes_atomique(onglet, data)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ── CLI (humain + agents) ─────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Gestionnaire du classeur de traçabilité Charmaz"
    )
    parser.add_argument("--init", action="store_true",
                        help="Créer tracabilite.xlsx avec les 11 onglets vides")
    parser.add_argument("--check", action="store_true",
                        help="Reconstruire checkpoint.json depuis le classeur")
    parser.add_argument("--status", action="store_true",
                        help="Afficher l'état courant")
    parser.add_argument("--list-sheets", action="store_true",
                        help="Liste les onglets avec leur nombre de lignes")
    parser.add_argument("--read-sheet", type=str, metavar="NAME",
                        help="Lit un onglet et affiche en JSON (ex: 01_Segments)")
    parser.add_argument("--limit", type=int, default=0,
                        help="Nombre max de lignes (0 = toutes)")
    parser.add_argument("--offset", type=int, default=0,
                        help="Nombre de lignes à sauter")
    parser.add_argument("--write-sheet", type=str, metavar="NAME",
                        help="Écrit une ligne dans un onglet depuis --data JSON")
    parser.add_argument("--data", type=str, default="{}",
                        help="Données JSON pour --write-sheet (ex: '{\"id_segment\":\"SEG_0001\",...}')")
    parser.add_argument("--write-batch", type=str, metavar="NAME",
                        help="Écrit plusieurs lignes depuis un fichier JSON")
    parser.add_argument("--batch-file", type=str,
                        help="Chemin du fichier JSON pour --write-batch")
    parser.add_argument("--get-ids", type=str, metavar="NAME",
                        help="Liste tous les IDs (1ère colonne) d'un onglet")
    parser.add_argument("--next-id", type=str, metavar="NAME",
                        help="Calcule le prochain ID disponible pour un onglet (ex: 02_Codage_Initial -> CODE_NNNN)")
    args = parser.parse_args()

    if args.init:
        creer_classeur()
    elif args.check:
        reconstruire_checkpoint()
    elif args.status:
        afficher_statut()
    elif args.list_sheets:
        if not XLSX_PATH.exists():
            print("Classeur introuvable. Lancer --init d'abord.")
            return
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if _[0])
            print(f"{name}: {count} lignes")
        wb.close()
    elif args.read_sheet:
        rows = lire_onglet_json(args.read_sheet, args.limit, args.offset)
        import json
        print(json.dumps(rows, ensure_ascii=False, indent=2))
    elif args.write_sheet:
        import json as _json
        data = _json.loads(args.data)
        n = ecrire_ligne_json(args.write_sheet, data)
        if n:
            id_val = data.get(list(data.keys())[0], "?")
            print(f"OK: ligne {id_val} écrite dans {args.write_sheet}")
        else:
            print(f"INFO: id déjà présent ou erreur")
    elif args.write_batch:
        if not args.batch_file:
            print("ERREUR: --batch-file requis avec --write-batch")
            return
        import json as _json
        with open(args.batch_file, encoding="utf-8") as f:
            lignes = _json.load(f)
        n = ecrire_lignes_depuis_json(args.write_batch, lignes)
        print(f"OK: {n} lignes écrites dans {args.write_batch}")
    elif args.get_ids:
        ids = lire_derniers_ids(args.get_ids)
        for i in sorted(ids):
            print(i)
        print(f"Total: {len(ids)} ids")
    elif args.next_id:
        prefix_map = {
            "00_Corpus_Source": "ENT_",
            "01_Segments": "SEG_",
            "02_Codage_Initial": "CODE_",
            "03_Codage_Focalise": "FCO_",
            "04_Categories": "CAT_",
            "05_Comparaisons": "CMP_",
            "06_Memos": "MEMO_",
            "07_Echantillonnage_Theorique": "ECH_",
            "08_Construction_Themes": "THM_",
            "09_Theorie_Finale": "ENO_",
            "10_Journal_Tracabilite": "DEC_",
        }
        prefix = prefix_map.get(args.next_id, "ID_")
        ids = lire_derniers_ids(args.next_id)
        nums = [int(i.split("_")[1]) for i in ids if i.startswith(prefix) and i.split("_")[1].isdigit()]
        next_num = (max(nums) + 1) if nums else 1
        if args.next_id in ("08_Construction_Themes",):
            # Sub-themes use STHM_ prefix
            sthm_nums = [int(i.split("_")[1]) for i in ids if i.startswith("STHM_") and i.split("_")[1].isdigit()]
            next_sthm = (max(sthm_nums) + 1) if sthm_nums else 1
            print(f"{prefix}{next_num:04d} / STHM_{next_sthm:04d}")
        else:
            fmt = "05d" if args.next_id in ("01_Segments", "02_Codage_Initial") else "04d"
            print(f"{prefix}{next_num:{fmt}}")
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
